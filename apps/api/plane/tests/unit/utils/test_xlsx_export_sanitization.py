# Copyright (c) 2023-present Plane Software, Inc. and contributors
# SPDX-License-Identifier: AGPL-3.0-only
# See the LICENSE file for details.

from io import BytesIO

import pytest
from openpyxl import load_workbook

from plane.utils.exporters.formatters import XLSXFormatter as SchemaXLSXFormatter
from plane.utils.porters.formatters import XLSXFormatter as PorterXLSXFormatter

# Characters that trigger formula evaluation in spreadsheet applications.
# See: https://owasp.org/www-community/attacks/CSV_Injection
FORMULA_TRIGGERS = ["=", "+", "-", "@", "\t", "\r", "\n"]

HYPERLINK_PAYLOAD = '=HYPERLINK("https://example.com/poc","click")'


def _load_cells(xlsx_bytes):
    """Load XLSX bytes and return the active worksheet's cells (formulas not evaluated)."""
    wb = load_workbook(filename=BytesIO(xlsx_bytes))
    return wb.active


@pytest.mark.unit
class TestPorterXLSXFormatterSanitization:
    """XLSX issue exports must not store user-controlled values as formula cells."""

    def test_formula_payload_is_stored_as_text(self):
        content = PorterXLSXFormatter().encode([{"name": HYPERLINK_PAYLOAD}])
        ws = _load_cells(content)
        cell = ws.cell(row=2, column=1)
        assert cell.data_type != "f"
        assert cell.value == "'" + HYPERLINK_PAYLOAD

    @pytest.mark.parametrize("trigger", FORMULA_TRIGGERS)
    def test_all_formula_trigger_characters_are_escaped(self, trigger):
        payload = trigger + "1+2"
        content = PorterXLSXFormatter().encode([{"name": payload}])
        ws = _load_cells(content)
        cell = ws.cell(row=2, column=1)
        assert cell.data_type != "f"
        assert cell.value == "'" + payload

    def test_safe_string_is_unchanged(self):
        content = PorterXLSXFormatter().encode([{"name": "Fix login bug"}])
        ws = _load_cells(content)
        assert ws.cell(row=2, column=1).value == "Fix login bug"

    def test_non_string_values_are_preserved(self):
        content = PorterXLSXFormatter().encode([{"estimate": 5}])
        ws = _load_cells(content)
        cell = ws.cell(row=2, column=1)
        assert cell.value == 5
        assert cell.data_type == "n"

    def test_list_value_joining_to_formula_is_escaped(self):
        content = PorterXLSXFormatter().encode([{"labels": ["=cmd", "bug"]}])
        ws = _load_cells(content)
        cell = ws.cell(row=2, column=1)
        assert cell.data_type != "f"
        assert cell.value == "'=cmd, bug"

    def test_headers_are_sanitized(self):
        content = PorterXLSXFormatter().encode([{"=evil_header": "value"}])
        ws = _load_cells(content)
        header = ws.cell(row=1, column=1)
        assert header.data_type != "f"
        assert header.value == "'=Evil Header"


class _FakeField:
    def __init__(self, label=None):
        self.label = label


class _FakeSchema:
    _declared_fields = {
        "name": _FakeField("Name"),
        "estimate": _FakeField("=Estimate"),
    }


@pytest.mark.unit
class TestSchemaXLSXFormatterSanitization:
    """Schema-based XLSX exports must not store user-controlled values as formula cells."""

    def test_formula_payload_is_stored_as_text(self):
        _, content = SchemaXLSXFormatter().format("export", [{"name": HYPERLINK_PAYLOAD, "estimate": 5}], _FakeSchema)
        ws = _load_cells(content)
        cell = ws.cell(row=2, column=1)
        assert cell.data_type != "f"
        assert cell.value == "'" + HYPERLINK_PAYLOAD

    @pytest.mark.parametrize("trigger", FORMULA_TRIGGERS)
    def test_all_formula_trigger_characters_are_escaped(self, trigger):
        payload = trigger + "1+2"
        _, content = SchemaXLSXFormatter().format("export", [{"name": payload, "estimate": 5}], _FakeSchema)
        ws = _load_cells(content)
        cell = ws.cell(row=2, column=1)
        assert cell.data_type != "f"
        assert cell.value == "'" + payload

    def test_safe_string_is_unchanged(self):
        _, content = SchemaXLSXFormatter().format("export", [{"name": "Fix login bug", "estimate": 5}], _FakeSchema)
        ws = _load_cells(content)
        assert ws.cell(row=2, column=1).value == "Fix login bug"

    def test_headers_are_sanitized(self):
        _, content = SchemaXLSXFormatter().format("export", [{"name": "ok", "estimate": 5}], _FakeSchema)
        ws = _load_cells(content)
        header = ws.cell(row=1, column=2)
        assert header.data_type != "f"
        assert header.value == "'=Estimate"
